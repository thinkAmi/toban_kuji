import unittest

import openpyxl

from make_data import FILE_NAME, USER_COUNT, GROUP_COUNT
from optimization import divide_users


class TestOptimization(unittest.TestCase):
    def create_ok_list(self):
        ok_list = {i: [] for i in range(18)}

        wb = openpyxl.load_workbook(FILE_NAME)
        ws = wb.worksheets[0]

        for col_index in range(2, USER_COUNT + 2):
            user_name = ws.cell(row=1, column=col_index).value

            for row_index in range(2, GROUP_COUNT + 2):
                if ws.cell(row=row_index, column=col_index).value == 1:
                    ok_list[row_index - 2].append(user_name)

        return ok_list

    def test_divide_users(self):
        group = divide_users()

        user_set = set()
        for users in group:
            # 1当番3名か
            self.assertEqual(len(users), 3)

            # 1回のみ割り当てられているか
            for user in users:
                self.assertNotIn(user, user_set)
                user_set.add(user)

        # 自分の希望した場所のみか
        ok_list = self.create_ok_list()
        for i, users in enumerate(group):
            for user in users:
                self.assertIn(user, ok_list[i])


if __name__ == '__main__':
    unittest.main()
