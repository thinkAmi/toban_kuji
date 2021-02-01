import random

import openpyxl


FILE_NAME = 'users_for_shuffle.xlsx'
USER_SHEET = 'users'
SCHEDULE_SHEET = 'schedule'

START_ROW = 2
END_ROW = 55
GROUP_COUNT = 18

GROUP_COL = 'A'


def read_users():
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb[USER_SHEET]
    users = []

    for i in range(START_ROW, END_ROW + 1):
        user_cell = f'A{i}'
        users.append(ws[user_cell].value)

    return users


def divide_users(users):
    current_index = 0

    # 当番の人のリストを要素として持つ、行事のリスト
    groups = [[] for _ in range(GROUP_COUNT)]

    random.shuffle(users)

    for user in users:
        groups[current_index].append(user)

        if current_index < GROUP_COUNT - 1:
            current_index += 1
        else:
            current_index = 0

    return groups


def write_users(groups):
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb[SCHEDULE_SHEET]

    # openpyxlの場合、セルのindexは1始まり
    for row_index, group in enumerate(groups, 2):  # 2行目から
        for col_index, user in enumerate(group, 2):  # B列目から
            ws.cell(row=row_index, column=col_index).value = user

    wb.save(FILE_NAME)


def main():
    users = read_users()
    groups = divide_users(users)
    write_users(groups)


if __name__ == '__main__':
    main()
