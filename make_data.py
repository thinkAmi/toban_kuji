import copy
import random

import openpyxl

OK = [1] * 10  # 18回中6割がOK
NG = [0] * 8

FILE_NAME = 'users_for_optimization.xlsx'
GROUP_COUNT = 18
USER_COUNT = 54


def main():
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    # タイトル行
    name_cell = ws.cell(row=1, column=1)
    name_cell.value = '開催番号'

    for i in range(2, USER_COUNT + 2):
        name_cell = ws.cell(row=1, column=i)
        name_cell.value = f'ユーザ_{i - 1}'

    # 開催回数列
    for i in range(2, GROUP_COUNT + 2):
        group_cell = ws.cell(row=i, column=1)
        group_cell.value = i - 2

    # 54人分のデータ
    for i in range(2, USER_COUNT + 2):
        ok = copy.deepcopy(OK)
        ng = copy.deepcopy(NG)
        total = ok + ng
        random.shuffle(total)

        # 18回の情報を埋める
        row = 2
        while total:
            result = total.pop()
            cell = ws.cell(row=row, column=i)
            cell.value = result

            row += 1

    wb.save(FILE_NAME)


if __name__ == "__main__":
    main()
