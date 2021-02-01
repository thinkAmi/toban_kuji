import numpy as np
import openpyxl
import pandas as pd
from ortoolpy import addbinvars, addvars
from pulp import LpProblem, lpSum, value, lpDot

from make_data import FILE_NAME, GROUP_COUNT, USER_COUNT

RESULT_FILE = 'result_optimization.xlsx'


def divide_users():
    # デフォルトだと 54 x 54 で読まれるので、不要行は読み込まないようにする
    skip_rows = [i + GROUP_COUNT for i in range(1, USER_COUNT - GROUP_COUNT + 1)]
    df = pd.read_excel(FILE_NAME, header=0, index_col=0, skiprows=skip_rows)

    # 当番回数
    event_count = df.shape[0]
    # print(f'{type(box_size)}: {box_size}')
    # => <class 'int'>: 18

    # ユーザ数
    user_count = df.shape[1]
    # print(f'{type(user_size)}: {user_size}')
    # => <class 'int'>: 54

    # 数理モデル
    model = LpProblem()

    # 変数を準備(当番/非当番の2値なので、0-1変数リスト)
    # https://docs.pyq.jp/python/math_opt/pdopt.html
    var_schedule = np.array(addbinvars(event_count, user_count))
    df['必要人数差'] = addvars(event_count)

    # 重み
    weight = 1

    # 目的関数の割り当て
    model += lpSum(df.必要人数差) * weight

    # 制約
    # 1当番あたり3人
    for idx, row in df.iterrows():
        model += row.必要人数差 >= (lpSum(var_schedule[row.name]) - 3)
        model += row.必要人数差 >= -(lpSum(var_schedule[row.name]) - 3)

    # 一人あたり1回当番すればよい
    for user in range(user_count):
        scheduled = [var_schedule[event, user] for event in range(event_count)]
        model += lpSum(pd.Series(scheduled)) <= 1

    # 当番可能な時だけ割り当てる
    df_rev = df[df.columns].apply(lambda r: 1 - r[df.columns], 1)
    for (_, d), (_, s) in zip(df_rev.iterrows(), pd.DataFrame(var_schedule).iterrows()):
        model += lpDot(d, s) <= 0

    # 実行
    model.solve()

    # 結果取得
    vectorized_results = np.vectorize(value)(var_schedule).astype(int)
    # print(type(vectorized_results))
    # => <class 'numpy.ndarray'>

    group = [[] for _ in range(event_count)]
    for i, vectorized_result in enumerate(vectorized_results):
        for result, name in zip(vectorized_result, df.columns):
            if result * name:
                group[i].append(name)

    return group


def export(group):
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    ws.cell(row=1, column=1).value = '開催番号'
    for i in range(2, 3 + 2):
        ws.cell(row=1, column=i).value = f'当番{i - 1}'

    for row_index, users in enumerate(group, 2):
        ws.cell(row=row_index, column=1).value = row_index - 2
        for col_index, user in enumerate(users, 2):
            ws.cell(row=row_index, column=col_index).value = user

    wb.save(RESULT_FILE)


def main():
    group = divide_users()
    export(group)


if __name__ == '__main__':
    main()
